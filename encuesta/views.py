from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.core.paginator import Paginator

from django.contrib.auth import authenticate, login, logout

from django.contrib import messages

from openpyxl import Workbook
from django.views.generic.base import TemplateView




from .models import Poll, Sig, Mejora

def inicio(request):

    context={}

    return render(request, 'poll/inicio.html', context)



def home(request, poll_id):
    
    """
    paginator= Paginator(polls, 1)
    pagina=request.GET.get("page") or 1
    polls=paginator.get_page(pagina)
    pagina_actual=int(pagina)
    paginas=range(1, polls.paginator.num_pages+1)
    """
    polls= Poll.objects.all()
    todas=Poll.objects.all().count()
    poll=Poll.objects.get(pk=poll_id)

    

    if request.method == 'POST':
   

        selected_option = request.POST['poll']
        if selected_option == 'option1':
            poll.option_one_count += 1
        elif selected_option == 'option2':
            poll.option_two_count += 1
        elif selected_option == 'option3':
            poll.option_three_count += 1
        else:
            return HttpResponse(400, 'Invalid form')
        
        


        poll.save()


        if poll.id < todas:
            return redirect('home', poll.id+1)
        else:
            return redirect ('sig')


    context = {
        'poll': poll,
        'polls': polls,
      
    }
    return render(request, 'poll/home.html', context)



def sig(request):  

      

    if request.method == 'POST': 

        comentario=Sig()      

        comentario.texto = request.POST['texto']

        comentario.save()
       
        return redirect ('mejora')

       
 
    context = {
     
      
      
    }
    return render(request, 'poll/sig.html', context)


def mejora(request):  

      

    if request.method == 'POST': 

        comentario=Mejora()      

        comentario.texto = request.POST['texto']

        comentario.save()
       
        return redirect ('gracias')

       
 
    context = {
     
      
      
    }
    return render(request, 'poll/mejora.html', context)


def gracias(request):
    
        return render(request, 'poll/gracias.html')


def results(request, poll_id):
    poll = Poll.objects.get(pk=poll_id)
    context = {
        'poll' : poll
    }
    return render(request, 'poll/results.html', context)

def resultados(request):
    polls = Poll.objects.all()
    context = {
        'polls' : polls
    }
    return render(request, 'poll/resultados.html', context)

def loginPage(request):

    if request.method=='POST':
        username=request.POST.get('username')
        password=request.POST.get('password')

        user=authenticate(request, username=username, password=password)

        if user:
            login(request, user)
            return  redirect('resultados')
        else:
            messages.info(request, 'Username or password es incorrecto No tienes autorizacion')

 

    return render(request, 'poll/login.html')

def logoutUser(request):
    logout(request)
    return redirect('inicio')

def ReporteExcel(request):

   
        polls= Poll.objects.all()
        wb=Workbook()
        ws=wb.active

        ws['B1']='REPORTE DE ENCUESTAS'

        ws.merge_cells('B1:G1')
        ws['B3']='Pregunta nº'
        ws['C3']='Enunciado'
        ws['D3']='SI'
        ws['E3']='NO'
        ws['F3']='NS/NC'
        ws['G3']='Total'

        cont=4

        for encuesta in polls:
            ws.cell(row=cont, column=2).value=encuesta.id
            ws.cell(row=cont, column=3).value=encuesta.question
            ws.cell(row=cont, column=4).value=encuesta.option_one_count
            ws.cell(row=cont, column=5).value=encuesta.option_two_count
            ws.cell(row=cont, column=6).value=encuesta.option_three_count
            ws.cell(row=cont, column=7).value=encuesta.option_one_count+encuesta.option_two_count+encuesta.option_three_count
            cont+=1

        nombre_archivo="ReporteEncuestas.xlsx"
        response=HttpResponse(content_type="application/ms-excel")
        content="attachment; filename={0}".format(nombre_archivo)
        response['Content-Disposition']=content
        wb.save(response)
        return response


